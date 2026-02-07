import json
import os

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side

from . import jobcan

class ExcelApp:
    def __init__(self, path, **kwargs):
        self.path = path
        self.modified = False

        if "override" in kwargs:
            override = kwargs["override"]
        else:
            override = False

        if override:
            self.wb = openpyxl.Workbook()
        elif os.path.isfile(path):
            self.wb = openpyxl.load_workbook(path)
        else:
            self.wb = openpyxl.Workbook()

        if "sheet" in kwargs:
            self.ws = self.wb[kwargs["sheet"]]
        else:
            self.ws = self.wb.active

        if "style" in kwargs:
            side = Side(style="thin", color="000000")
            border = Border(bottom=side)
            col = 1
            for f in kwargs["style"]:
                self.ws.column_dimensions[f[0]].width = f[2]
                self.ws.cell(1, col).value = f[1]
                self.ws.cell(1, col).border = border
                col += 1

    def __del__(self):
        if self.modified:
            self.wb.save(self.path)

    def add_image(self, img_path, cell):
        img = Image(img_path)
        self.ws.add_image(img, cell)

    def save(self):
        if self.modified:
            self.wb.save(self.path)

    def write(self, row, col, value):
        self.modified = True
        self.ws.cell(row, col).value = value

    def number_format(self, row, col, number_format):
        self.modified = True
        self.ws.cell(row, col).number_format = number_format

    def hyperlink(self, row, col, link):
        self.modified = True
        self.ws.cell(row, col).hyperlink = link

    def value(self, row, col):
        return self.ws.cell(row, col).value

    def find_empty_row(self):
        row = 1
        while self.ws.cell(row, 1).value is not None:
            row += 1

        return row

    def format_cell(self, row, col, fmt, jobcan_id):
        if "format" in fmt:
            self.number_format(row, col, fmt["format"])
        if "link" in fmt and fmt["link"] == "request":
            self.hyperlink(row, col, f"https://ssl.wf.jobcan.jp/#/requests/{jobcan_id}")

def excel_write_detail(ex, details, detail_format, row, col):
    r = 0
    for detail in details:
        c = 0
        step_col = len(detail) 
        for d in detail:
            if detail_format[c]["format_type"] == "master":
                if type(d['value']) == dict:
                    ex.write(row + r, col + c, d['value']['record_name'])
                else:
                    ex.write(row + r, col + c, d['value'])
            elif detail_format[c]["format_type"] == "price":
                ex.write(row + r, col + c, d['value'])
                ex.number_format(row + r, col + c, detail_format[c]["format"])
            else:
                ex.write(row + r, col + c, d['value'])
                
            c += 1
        r += 1
    step_row = len(details)

    return(step_row, step_col)
                
    
def excel(request_list, args, output_format, excel_style, **kwargs):
    ex = ExcelApp(args["output_file"], style=excel_style, override=args["override"])

    if 'detail_format' in kwargs:
        detail_format = kwargs['detail_format']
    else:
        detail_format = None

    row = 2
    for item in request_list:
        item["applicant_full_name"] = (
            item["applicant_last_name"] + " " + item["applicant_first_name"]
        )
        col = 1
        step_row = 1
        for fmt in output_format:
            step_col = 1
            tmp = ""
            if fmt["column"] == "custom":
                ci = item["detail"]["customized_items"]
                tmp = jobcan.parse_customized_items(ci, fmt["custom"], "content")
                ex.write(row, col, tmp)
                ex.format_cell(row, col, fmt, tmp)
            elif fmt["column"] == "detail":
                ci = item["detail"]["customized_items"]
                details = jobcan.parse_customized_items(ci, "明細", "table")
                step_row, step_col = excel_write_detail(ex, details, detail_format, row, col)
            else:
                tmp = item[fmt["column"]]
                ex.write(row, col, tmp)
                ex.format_cell(row, col, fmt, tmp)

            col += step_col
        row += step_row


def console(json_dict):
    print(json.dumps(json_dict, indent=2, ensure_ascii=False))
