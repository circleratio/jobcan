import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.drawing.image import Image
import os
import json


class excel:
    def __init__(self, path, **kwargs):
        self.path = path
        self.modified = False

        if "override" in kwargs:
            override = kwargs["override"]
        else:
            override = False

        if override:
            self.wb = openpyxl.Workbook()
        elif os.path.isfile(path):
            self.wb = openpyxl.load_workbook(path)
        else:
            self.wb = openpyxl.Workbook()

        if "sheet" in kwargs:
            self.ws = self.wb[kwargs["sheet"]]
        else:
            self.ws = self.wb.active

        if "style" in kwargs:
            side = Side(style="thin", color="000000")
            border = Border(bottom=side)
            col = 1
            for f in kwargs["style"]:
                self.ws.column_dimensions[f[0]].width = f[2]
                self.ws.cell(1, col).value = f[1]
                self.ws.cell(1, col).border = border
                col += 1

    def __del__(self):
        if self.modified:
            self.wb.save(self.path)

    def add_image(self, img_path, cell):
        img = Image(img_path)
        self.ws.add_image(img, cell)

    def save(self):
        if self.modified:
            self.wb.save(self.path)

    def write(self, row, col, value):
        self.modified = True
        self.ws.cell(row, col).value = value

    def number_format(self, row, col, number_format):
        self.modified = True
        self.ws.cell(row, col).number_format = number_format

    def value(self, row, col):
        return self.ws.cell(row, col).value

    def find_empty_row(self):
        row = 1
        while self.ws.cell(row, 1).value is not None:
            row += 1

        return row


def console(json_dict):
    print(json.dumps(json_dict, indent=2, ensure_ascii=False))
